# openpyxl - ler e alterar dados de um arquivo Excel
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregar excel existente
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome da planilha
sheet_name = 'Planilha1'

# Selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell] 
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 23)
    print()

worksheet['B3'].value = 26

workbook.save(WORKBOOK_PATH)
