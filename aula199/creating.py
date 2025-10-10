# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()

# Nome da planilha
sheet_name = 'Planilha1'

# Criar uma nova planilha
workbook.create_sheet(sheet_name, 0)

# Selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover a planilha padrão
workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1,1,'Nome')
worksheet.cell(1,2,'Idade')
worksheet.cell(1,3,'Nota')

students = [
    # nome  idade , media
    ['Fernando', 29, 8.5],
    ['Ana', 25, 9.5],
    ['João', 30, 6.5],
    ['Maria', 22, 7.5],
    ['Pedro', 28, 5.5],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for studend in students:
    worksheet.append(studend)

workbook.save(WORKBOOK_PATH)
